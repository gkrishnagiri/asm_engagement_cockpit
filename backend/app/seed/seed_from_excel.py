from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.database import SessionLocal, create_database_tables
from app.models import Deliverable, Engagement, Task, Workstream

DEFAULT_MY_WORK_OWNER = "Giridhar"
DEFAULT_EXCEL_PATH = Path("/home/induser1/giri/AIProjects/asm_engagement_cockpit/data/initial_tracker.xlsx")

ENGAGEMENT_NAME = "Mondelez AMS 2.0 Internal Delivery Tracker"
CLIENT_NAME = "Mondelez"

OWNERSHIP_TABLES = ("workstreams", "deliverables", "tasks", "subtasks")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_compact(value: Any) -> str:
    return " ".join(normalize_text(value).lower().split())


def contains_owner(value: Any, owner_name: str = DEFAULT_MY_WORK_OWNER) -> bool:
    return normalize_compact(owner_name) in normalize_compact(value)


def excel_serial_to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()

    text_value = normalize_text(value)
    if not text_value:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue

    return None


def normalize_workstream_external_id(value: Any, workstream_name: str | None = None) -> str:
    raw_value = normalize_text(value)

    if raw_value == "-":
        return "WS8"

    if raw_value.upper().startswith("WS"):
        return raw_value.split()[0].strip().upper()

    normalized_name = normalize_text(workstream_name).lower()

    if "automation" in normalized_name and "roadmap" in normalized_name:
        return "WS8"

    return raw_value.upper()


def normalize_deliverable_external_id(value: Any) -> str:
    return normalize_text(value).upper()


def get_model_columns(model: type[Any]) -> set[str]:
    return {column.key for column in model.__table__.columns}


def set_if_column(item: Any, column_name: str, value: Any) -> None:
    if column_name in get_model_columns(type(item)):
        setattr(item, column_name, value)


def update_owner_columns(session: Session, table_name: str, entity_id: Any, owner_name: str | None, secondary_owner_name: str | None) -> None:
    if table_name not in OWNERSHIP_TABLES:
        return

    session.execute(
        text(
            f"""
            UPDATE {table_name}
            SET owner_name = :owner_name,
                secondary_owner_name = :secondary_owner_name
            WHERE id = :entity_id
            """
        ),
        {
            "owner_name": owner_name,
            "secondary_owner_name": secondary_owner_name,
            "entity_id": str(entity_id),
        },
    )


def ensure_ownership_columns(session: Session) -> None:
    for table_name in OWNERSHIP_TABLES:
        session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS owner_name VARCHAR(250)"))
        session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS secondary_owner_name VARCHAR(500)"))
    session.commit()


def read_sheet_rows(workbook, sheet_name: str, header_row_number: int) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [normalize_text(cell.value) for cell in sheet[header_row_number]]

    rows: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        row_dict = {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
            if headers[index]
        }

        if any(value is not None and normalize_text(value) for value in row_dict.values()):
            rows.append(row_dict)

    return rows


def get_or_create_engagement(session: Session) -> Engagement:
    existing = session.scalar(select(Engagement).where(Engagement.name == ENGAGEMENT_NAME))
    if existing is not None:
        existing.client_name = CLIENT_NAME
        existing.description = (
            "Seeded from the Mondelez AMS 2.0 HCLTech Internal Delivery Tracker. "
            "The seed now loads all valid tracker rows; My Work filtering is handled by the UI/API."
        )
        existing.status = existing.status or "In Progress"
        return existing

    engagement = Engagement(
        name=ENGAGEMENT_NAME,
        client_name=CLIENT_NAME,
        description=(
            "Seeded from the Mondelez AMS 2.0 HCLTech Internal Delivery Tracker. "
            "The seed now loads all valid tracker rows; My Work filtering is handled by the UI/API."
        ),
        status="In Progress",
    )

    session.add(engagement)
    session.flush()

    return engagement


def seed_workstreams(session: Session, workbook, engagement: Engagement) -> dict[str, Workstream]:
    rows = read_sheet_rows(workbook, "Workstream Ownership", 1)
    workstreams_by_external_id: dict[str, Workstream] = {}

    for row in rows:
        external_id = normalize_workstream_external_id(row.get("WS"), row.get("Workstream"))
        name = normalize_text(row.get("Workstream"))

        if not external_id or not name:
            continue

        primary_owner = normalize_text(row.get("Primary HCL Owner")) or None
        support_owner = normalize_text(row.get("Supporting Team")) or None

        existing = session.scalar(
            select(Workstream).where(
                Workstream.engagement_id == engagement.id,
                Workstream.external_id == external_id,
            )
        )

        if existing is None:
            item = Workstream(
                engagement_id=engagement.id,
                external_id=external_id,
                name=name,
                objective=normalize_text(row.get("Core Internal Responsibilities")) or None,
                scope=normalize_text(row.get("Committed Deliverables")) or None,
                target_completion_date=excel_serial_to_date(row.get("Internal Due / Gate (TBC)")),
                status="Not Started",
            )
            set_if_column(item, "owner_name", primary_owner)
            set_if_column(item, "secondary_owner_name", support_owner)
            session.add(item)
            session.flush()
            existing = item
        else:
            existing.name = name
            existing.objective = normalize_text(row.get("Core Internal Responsibilities")) or existing.objective
            existing.scope = normalize_text(row.get("Committed Deliverables")) or existing.scope
            existing.target_completion_date = excel_serial_to_date(row.get("Internal Due / Gate (TBC)")) or existing.target_completion_date
            set_if_column(existing, "owner_name", primary_owner)
            set_if_column(existing, "secondary_owner_name", support_owner)

        update_owner_columns(session, "workstreams", existing.id, primary_owner, support_owner)
        workstreams_by_external_id[external_id] = existing

    return workstreams_by_external_id


def get_or_create_placeholder_workstream(session: Session, engagement: Engagement, workstreams_by_external_id: dict[str, Workstream], external_id: str) -> Workstream:
    existing = workstreams_by_external_id.get(external_id)
    if existing is not None:
        return existing

    existing = session.scalar(
        select(Workstream).where(
            Workstream.engagement_id == engagement.id,
            Workstream.external_id == external_id,
        )
    )
    if existing is not None:
        workstreams_by_external_id[external_id] = existing
        return existing

    item = Workstream(
        engagement_id=engagement.id,
        external_id=external_id,
        name=f"{external_id} Workstream",
        status="Not Started",
    )
    session.add(item)
    session.flush()
    update_owner_columns(session, "workstreams", item.id, None, None)
    workstreams_by_external_id[external_id] = item
    return item


def seed_deliverables(session: Session, workbook, engagement: Engagement, workstreams_by_external_id: dict[str, Workstream]) -> dict[str, Deliverable]:
    rows = read_sheet_rows(workbook, "Deliverable Control", 4)
    deliverables_by_external_id: dict[str, Deliverable] = {}

    for row in rows:
        deliverable_external_id = normalize_deliverable_external_id(row.get("ID") or row.get("ID2"))
        workstream_external_id = normalize_workstream_external_id(row.get("Workstream"), row.get("Workstream"))
        name = normalize_text(row.get("Deliverable"))

        if not deliverable_external_id or not workstream_external_id or not name:
            continue

        primary_owner = normalize_text(row.get("Primary Owner")) or None
        support_owner = normalize_text(row.get("Support")) or None

        workstream = get_or_create_placeholder_workstream(session, engagement, workstreams_by_external_id, workstream_external_id)

        existing = session.scalar(
            select(Deliverable).where(
                Deliverable.workstream_id == workstream.id,
                Deliverable.external_id == deliverable_external_id,
            )
        )

        if existing is None:
            item = Deliverable(
                workstream_id=workstream.id,
                external_id=deliverable_external_id,
                name=name,
                description=normalize_text(row.get("Acceptance / Completion Evidence")) or None,
                deliverable_type="Consulting Deliverable",
                start_date=excel_serial_to_date(row.get("Draft Due (TBC)")),
                target_completion_date=excel_serial_to_date(row.get("Final Due (TBC)")),
                status=normalize_text(row.get("Status")) or "Not Started",
                review_status="Not Submitted",
            )
            set_if_column(item, "owner_name", primary_owner)
            set_if_column(item, "secondary_owner_name", support_owner)
            session.add(item)
            session.flush()
            existing = item
        else:
            existing.name = name
            existing.description = normalize_text(row.get("Acceptance / Completion Evidence")) or existing.description
            existing.start_date = excel_serial_to_date(row.get("Draft Due (TBC)")) or existing.start_date
            existing.target_completion_date = excel_serial_to_date(row.get("Final Due (TBC)")) or existing.target_completion_date
            existing.status = normalize_text(row.get("Status")) or existing.status
            set_if_column(existing, "owner_name", primary_owner)
            set_if_column(existing, "secondary_owner_name", support_owner)

        update_owner_columns(session, "deliverables", existing.id, primary_owner, support_owner)
        deliverables_by_external_id[deliverable_external_id] = existing

    return deliverables_by_external_id


def find_deliverable_by_name_and_workstream(
    session: Session,
    deliverable_name: str,
    workstream_external_id: str,
    workstreams_by_external_id: dict[str, Workstream],
) -> Deliverable | None:
    if not deliverable_name or not workstream_external_id:
        return None

    workstream = workstreams_by_external_id.get(workstream_external_id)
    if workstream is None:
        return None

    normalized_target = normalize_compact(deliverable_name)
    candidates = list(session.scalars(select(Deliverable).where(Deliverable.workstream_id == workstream.id)).all())

    for candidate in candidates:
        if normalize_compact(candidate.name) == normalized_target:
            return candidate

    for candidate in candidates:
        if normalized_target and normalized_target in normalize_compact(candidate.name):
            return candidate

    return None


def get_or_create_deliverable_for_task(
    session: Session,
    row: dict[str, Any],
    engagement: Engagement,
    workstreams_by_external_id: dict[str, Workstream],
    deliverables_by_external_id: dict[str, Deliverable],
) -> Deliverable | None:
    deliverable_external_id = normalize_deliverable_external_id(row.get("Deliverable ID"))
    deliverable_name = normalize_text(row.get("Deliverable"))
    workstream_external_id = normalize_workstream_external_id(row.get("Workstream2") or row.get("Workstream"), row.get("Workstream"))

    if deliverable_external_id and deliverable_external_id in deliverables_by_external_id:
        return deliverables_by_external_id[deliverable_external_id]

    if deliverable_external_id:
        existing = session.scalar(select(Deliverable).where(Deliverable.external_id == deliverable_external_id))
        if existing is not None:
            deliverables_by_external_id[deliverable_external_id] = existing
            return existing

    by_name = find_deliverable_by_name_and_workstream(session, deliverable_name, workstream_external_id, workstreams_by_external_id)
    if by_name is not None:
        if by_name.external_id:
            deliverables_by_external_id[by_name.external_id] = by_name
        return by_name

    if not deliverable_external_id and not deliverable_name:
        return None

    workstream = get_or_create_placeholder_workstream(session, engagement, workstreams_by_external_id, workstream_external_id or "UNMAPPED")
    placeholder_id = deliverable_external_id or f"UNMAPPED-{normalize_text(row.get('Task ID')) or 'TASK'}"
    placeholder_name = deliverable_name or f"Unmapped Deliverable for {placeholder_id}"

    item = Deliverable(
        workstream_id=workstream.id,
        external_id=placeholder_id,
        name=placeholder_name,
        description="Placeholder deliverable created by seed loader because the task referenced a deliverable not found in Deliverable Control.",
        deliverable_type="Consulting Deliverable",
        status="Not Started",
        review_status="Not Submitted",
    )
    session.add(item)
    session.flush()
    update_owner_columns(session, "deliverables", item.id, None, None)
    deliverables_by_external_id[placeholder_id] = item
    return item


def seed_tasks(
    session: Session,
    workbook,
    engagement: Engagement,
    deliverables_by_external_id: dict[str, Deliverable],
    workstreams_by_external_id: dict[str, Workstream],
) -> dict[str, Task]:
    rows = read_sheet_rows(workbook, "Detailed Task Tracker", 1)
    tasks_by_external_id: dict[str, Task] = {}

    for row in rows:
        task_external_id = normalize_text(row.get("Task ID")).upper()
        title = normalize_text(row.get("Activity / Task"))

        if not task_external_id or not title:
            continue

        deliverable = get_or_create_deliverable_for_task(
            session,
            row,
            engagement,
            workstreams_by_external_id,
            deliverables_by_external_id,
        )

        if deliverable is None:
            continue

        primary_owner = normalize_text(row.get("Primary Owner")) or None
        support_owner = normalize_text(row.get("Support")) or None

        existing = session.scalar(
            select(Task).where(
                Task.deliverable_id == deliverable.id,
                Task.external_id == task_external_id,
            )
        )

        description_parts = [
            f"Workstream: {normalize_text(row.get('Workstream'))}",
            f"Workstream ID: {normalize_text(row.get('Workstream2'))}",
            f"Deliverable: {normalize_text(row.get('Deliverable'))}",
            f"Deliverable ID: {normalize_text(row.get('Deliverable ID'))}",
            f"Phase: {normalize_text(row.get('Phase'))}",
            f"Internal checkpoint: {normalize_text(row.get('Internal Checkpoint'))}",
            f"RAG: {normalize_text(row.get('RAG'))}",
            f"Dependency/Input Needed: {normalize_text(row.get('Dependency / Input Needed'))}",
            f"Next Action: {normalize_text(row.get('Next Action'))}",
            f"Due Health: {normalize_text(row.get('Due Health'))}",
            f"Support: {support_owner or ''}",
        ]
        description = "\n".join(part for part in description_parts if not part.endswith(": "))

        if existing is None:
            item = Task(
                deliverable_id=deliverable.id,
                external_id=task_external_id,
                title=title,
                description=description or None,
                priority=normalize_text(row.get("Priority")) or None,
                start_date=excel_serial_to_date(row.get("Start Date")),
                target_completion_date=excel_serial_to_date(row.get("Due Date")),
                status=normalize_text(row.get("Status")) or "Not Started",
            )
            set_if_column(item, "owner_name", primary_owner)
            set_if_column(item, "secondary_owner_name", support_owner)
            session.add(item)
            session.flush()
            existing = item
        else:
            existing.title = title
            existing.description = description or existing.description
            existing.priority = normalize_text(row.get("Priority")) or existing.priority
            existing.start_date = excel_serial_to_date(row.get("Start Date")) or existing.start_date
            existing.target_completion_date = excel_serial_to_date(row.get("Due Date")) or existing.target_completion_date
            existing.status = normalize_text(row.get("Status")) or existing.status
            set_if_column(existing, "owner_name", primary_owner)
            set_if_column(existing, "secondary_owner_name", support_owner)

        update_owner_columns(session, "tasks", existing.id, primary_owner, support_owner)
        tasks_by_external_id[task_external_id] = existing

    return tasks_by_external_id


def count_owned(session: Session, table_name: str, owner_column: str, owner_name: str = DEFAULT_MY_WORK_OWNER) -> int:
    return int(
        session.execute(
            text(
                f"""
                SELECT COUNT(*)
                FROM {table_name}
                WHERE lower(coalesce({owner_column}, '')) LIKE :owner_pattern
                """
            ),
            {"owner_pattern": f"%{owner_name.lower()}%"},
        ).scalar()
        or 0
    )


def run_seed(excel_path: Path) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel file not found: {excel_path}\n"
            "Save the tracker file as data/initial_tracker.xlsx or pass --file <path>."
        )

    create_database_tables()

    workbook = load_workbook(excel_path, data_only=True)

    with SessionLocal() as session:
        ensure_ownership_columns(session)

        engagement = get_or_create_engagement(session)

        workstreams_by_external_id = seed_workstreams(session, workbook, engagement)
        deliverables_by_external_id = seed_deliverables(
            session,
            workbook,
            engagement,
            workstreams_by_external_id,
        )
        tasks_by_external_id = seed_tasks(
            session,
            workbook,
            engagement,
            deliverables_by_external_id,
            workstreams_by_external_id,
        )

        session.commit()

        print("")
        print("Seed completed successfully.")
        print(f"Engagement:   {engagement.name}")
        print(f"Workstreams:  {len(workstreams_by_external_id)}")
        print(f"Deliverables: {len(deliverables_by_external_id)}")
        print(f"Tasks:        {len(tasks_by_external_id)}")
        print("")
        print("My Work Summary for Giridhar:")
        print(f"Primary workstreams:    {count_owned(session, 'workstreams', 'owner_name')}")
        print(f"Secondary workstreams:  {count_owned(session, 'workstreams', 'secondary_owner_name')}")
        print(f"Primary deliverables:   {count_owned(session, 'deliverables', 'owner_name')}")
        print(f"Secondary deliverables: {count_owned(session, 'deliverables', 'secondary_owner_name')}")
        print(f"Primary tasks:          {count_owned(session, 'tasks', 'owner_name')}")
        print(f"Secondary tasks:        {count_owned(session, 'tasks', 'secondary_owner_name')}")
        print("")
        print("Seed rule:")
        print("- All valid tracker rows were loaded; owner filtering is not applied during seed.")
        print("- Workstream ownership uses Primary HCL Owner and Supporting Team.")
        print("- Deliverable ownership uses Primary Owner and Support.")
        print("- Task ownership uses Primary Owner and Support.")
        print("- Tasks are mapped to deliverables using Deliverable ID from the Detailed Task Tracker.")
        print("- Workstream '-' for AI and Automation Roadmap Advisory is normalized to WS8.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Seed ASM Engagement Cockpit from Excel tracker.")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="Path to the Excel tracker file.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run_seed(args.file)
