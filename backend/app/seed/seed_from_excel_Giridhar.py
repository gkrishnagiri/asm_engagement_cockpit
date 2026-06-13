from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import SessionLocal, create_database_tables
from app.models import Deliverable, Engagement, Task, Workstream

PRIMARY_OWNER = "Giridhar Krishnagiri"
DEFAULT_EXCEL_PATH = Path("/home/induser1/giri/AIProjects/asm_engagement_cockpit/data/initial_tracker.xlsx")

ENGAGEMENT_NAME = "Mondelez AMS 2.0 Internal Delivery Tracker"
CLIENT_NAME = "Mondelez"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_owner(value: Any) -> str:
    return normalize_text(value).lower()


def is_primary_owner(value: Any) -> bool:
    return normalize_owner(value) == PRIMARY_OWNER.lower()


def excel_serial_to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, (int, float)):
        # Excel date serials use 1899-12-30 as the common conversion base.
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()

    text_value = str(value).strip()
    if not text_value:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue

    return None


def normalize_workstream_external_id(value: Any, workstream_name: str | None = None) -> str:
    raw_value = normalize_text(value)

    if raw_value == "-":
        return "WS8"

    if raw_value.startswith("WS"):
        return raw_value.split()[0].strip()

    normalized_name = normalize_text(workstream_name).lower()

    if "automation" in normalized_name and "roadmap" in normalized_name:
        return "WS8"

    return raw_value


def normalize_workstream_name_from_task(value: Any) -> tuple[str, str]:
    raw_value = normalize_text(value)

    if not raw_value:
        return "", ""

    parts = raw_value.split(" ", 1)

    if len(parts) == 1:
        external_id = normalize_workstream_external_id(parts[0], raw_value)
        return external_id, raw_value

    external_id = normalize_workstream_external_id(parts[0], raw_value)
    name = parts[1].strip()

    return external_id, name


def get_or_create_engagement(session) -> Engagement:
    existing = session.scalar(select(Engagement).where(Engagement.name == ENGAGEMENT_NAME))
    if existing is not None:
        return existing

    engagement = Engagement(
        name=ENGAGEMENT_NAME,
        client_name=CLIENT_NAME,
        description=(
            "Seeded from the Mondelez AMS 2.0 HCLTech Internal Delivery Tracker. "
            "Initial data includes primary-owned workstreams, deliverables, and tasks."
        ),
        status="In Progress",
    )

    session.add(engagement)
    session.flush()

    return engagement


def read_sheet_rows(workbook, sheet_name: str, header_row_number: int) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [normalize_text(cell.value) for cell in sheet[header_row_number]]

    rows: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        row_dict = {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
            if headers[index]
        }

        if any(value is not None and normalize_text(value) for value in row_dict.values()):
            rows.append(row_dict)

    return rows


def seed_workstreams(session, workbook, engagement: Engagement) -> dict[str, Workstream]:
    rows = read_sheet_rows(workbook, "Workstream Ownership", 1)
    workstreams_by_external_id: dict[str, Workstream] = {}

    for row in rows:
        if not is_primary_owner(row.get("Primary HCL Owner")):
            continue

        external_id = normalize_workstream_external_id(row.get("WS"), row.get("Workstream"))
        name = normalize_text(row.get("Workstream"))

        if not external_id or not name:
            continue

        existing = session.scalar(
            select(Workstream).where(
                Workstream.engagement_id == engagement.id,
                Workstream.external_id == external_id,
            )
        )

        if existing is None:
            item = Workstream(
                engagement_id=engagement.id,
                external_id=external_id,
                name=name,
                objective=normalize_text(row.get("Core Internal Responsibilities")) or None,
                scope=normalize_text(row.get("Committed Deliverables")) or None,
                target_completion_date=excel_serial_to_date(row.get("Internal Due / Gate (TBC)")),
                status="Not Started",
            )
            session.add(item)
            session.flush()
            existing = item
        else:
            existing.name = name
            existing.objective = normalize_text(row.get("Core Internal Responsibilities")) or existing.objective
            existing.scope = normalize_text(row.get("Committed Deliverables")) or existing.scope
            existing.target_completion_date = (
                excel_serial_to_date(row.get("Internal Due / Gate (TBC)"))
                or existing.target_completion_date
            )

        workstreams_by_external_id[external_id] = existing

    return workstreams_by_external_id


def seed_deliverables(
    session,
    workbook,
    engagement: Engagement,
    workstreams_by_external_id: dict[str, Workstream],
) -> dict[str, Deliverable]:
    rows = read_sheet_rows(workbook, "Deliverable Control", 4)
    deliverables_by_external_id: dict[str, Deliverable] = {}

    for row in rows:
        if not is_primary_owner(row.get("Primary Owner")):
            continue

        deliverable_external_id = normalize_text(row.get("ID"))
        workstream_external_id = normalize_workstream_external_id(row.get("Workstream"), row.get("Workstream"))
        name = normalize_text(row.get("Deliverable"))

        if not deliverable_external_id or not workstream_external_id or not name:
            continue

        workstream = workstreams_by_external_id.get(workstream_external_id)

        if workstream is None:
            workstream = session.scalar(
                select(Workstream).where(
                    Workstream.engagement_id == engagement.id,
                    Workstream.external_id == workstream_external_id,
                )
            )

        if workstream is None:
            workstream = Workstream(
                engagement_id=engagement.id,
                external_id=workstream_external_id,
                name=f"{workstream_external_id} Workstream",
                status="Not Started",
            )
            session.add(workstream)
            session.flush()
            workstreams_by_external_id[workstream_external_id] = workstream

        existing = session.scalar(
            select(Deliverable).where(
                Deliverable.workstream_id == workstream.id,
                Deliverable.external_id == deliverable_external_id,
            )
        )

        if existing is None:
            item = Deliverable(
                workstream_id=workstream.id,
                external_id=deliverable_external_id,
                name=name,
                description=normalize_text(row.get("Acceptance / Completion Evidence")) or None,
                deliverable_type="Consulting Deliverable",
                start_date=excel_serial_to_date(row.get("Draft Due (TBC)")),
                target_completion_date=excel_serial_to_date(row.get("Final Due (TBC)")),
                status=normalize_text(row.get("Status")) or "Not Started",
                review_status="Not Submitted",
            )
            session.add(item)
            session.flush()
            existing = item
        else:
            existing.name = name
            existing.description = normalize_text(row.get("Acceptance / Completion Evidence")) or existing.description
            existing.start_date = excel_serial_to_date(row.get("Draft Due (TBC)")) or existing.start_date
            existing.target_completion_date = excel_serial_to_date(row.get("Final Due (TBC)")) or existing.target_completion_date
            existing.status = normalize_text(row.get("Status")) or existing.status

        deliverables_by_external_id[deliverable_external_id] = existing

    return deliverables_by_external_id


def find_best_deliverable_for_task(
    task_row: dict[str, Any],
    deliverables_by_external_id: dict[str, Deliverable],
    workstreams_by_external_id: dict[str, Workstream],
) -> Deliverable | None:
    workstream_external_id, _ = normalize_workstream_name_from_task(task_row.get("Workstream"))

    candidate_deliverables = [
        deliverable
        for deliverable in deliverables_by_external_id.values()
        if workstreams_by_external_id.get(workstream_external_id) is not None
        and deliverable.workstream_id == workstreams_by_external_id[workstream_external_id].id
    ]

    if not candidate_deliverables:
        return None

    # MVP 2 seed rule:
    # Tasks in the Excel tracker are workstream-level execution tasks.
    # Until the UI supports explicit task-to-deliverable mapping, attach each task
    # to the first primary-owned deliverable in the same workstream.
    return sorted(candidate_deliverables, key=lambda item: item.external_id or "")[0]


def seed_tasks(
    session,
    workbook,
    deliverables_by_external_id: dict[str, Deliverable],
    workstreams_by_external_id: dict[str, Workstream],
) -> dict[str, Task]:
    rows = read_sheet_rows(workbook, "Detailed Task Tracker", 1)
    tasks_by_external_id: dict[str, Task] = {}

    for row in rows:
        if not is_primary_owner(row.get("Primary Owner")):
            continue

        task_external_id = normalize_text(row.get("Task ID"))
        title = normalize_text(row.get("Activity / Task"))

        if not task_external_id or not title:
            continue

        deliverable = find_best_deliverable_for_task(
            row,
            deliverables_by_external_id,
            workstreams_by_external_id,
        )

        if deliverable is None:
            continue

        existing = session.scalar(
            select(Task).where(
                Task.deliverable_id == deliverable.id,
                Task.external_id == task_external_id,
            )
        )

        description_parts = [
            f"Phase: {normalize_text(row.get('Phase'))}",
            f"Internal checkpoint: {normalize_text(row.get('Internal Checkpoint'))}",
            f"Dependency/Input Needed: {normalize_text(row.get('Dependency / Input Needed'))}",
            f"Next Action: {normalize_text(row.get('Next Action'))}",
            f"Due Health: {normalize_text(row.get('Due Health'))}",
            f"Support: {normalize_text(row.get('Support'))}",
        ]
        description = "\n".join(part for part in description_parts if not part.endswith(": "))

        if existing is None:
            item = Task(
                deliverable_id=deliverable.id,
                external_id=task_external_id,
                title=title,
                description=description or None,
                priority=normalize_text(row.get("Priority")) or None,
                start_date=excel_serial_to_date(row.get("Start Date")),
                target_completion_date=excel_serial_to_date(row.get("Due Date")),
                status=normalize_text(row.get("Status")) or "Not Started",
            )
            session.add(item)
            session.flush()
            existing = item
        else:
            existing.title = title
            existing.description = description or existing.description
            existing.priority = normalize_text(row.get("Priority")) or existing.priority
            existing.start_date = excel_serial_to_date(row.get("Start Date")) or existing.start_date
            existing.target_completion_date = excel_serial_to_date(row.get("Due Date")) or existing.target_completion_date
            existing.status = normalize_text(row.get("Status")) or existing.status

        tasks_by_external_id[task_external_id] = existing

    return tasks_by_external_id


def run_seed(excel_path: Path) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel file not found: {excel_path}\n"
            "Save the tracker file as data/initial_tracker.xlsx or pass --file <path>."
        )

    create_database_tables()

    workbook = load_workbook(excel_path, data_only=True)

    with SessionLocal() as session:
        engagement = get_or_create_engagement(session)

        workstreams_by_external_id = seed_workstreams(session, workbook, engagement)
        deliverables_by_external_id = seed_deliverables(
            session,
            workbook,
            engagement,
            workstreams_by_external_id,
        )
        tasks_by_external_id = seed_tasks(
            session,
            workbook,
            deliverables_by_external_id,
            workstreams_by_external_id,
        )

        session.commit()

        print("")
        print("Seed completed successfully.")
        print(f"Engagement:   {engagement.name}")
        print(f"Workstreams:  {len(workstreams_by_external_id)}")
        print(f"Deliverables: {len(deliverables_by_external_id)}")
        print(f"Tasks:        {len(tasks_by_external_id)}")
        print("")
        print("Seed rule:")
        print("- Only rows where Giridhar Krishnagiri is the primary owner were loaded.")
        print("- Workstream '-' for AI and Automation Roadmap Advisory was normalized to WS8.")
        print("- Tasks were attached to the first primary-owned deliverable in the matching workstream.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Seed ASM Engagement Cockpit from Excel tracker.")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="Path to the Excel tracker file.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run_seed(args.file)